"""Main parser orchestrator for the Slesh Event Plan Excel.

Public function: parse_event_plan(data: bytes | BinaryIO) -> ParsedEventPlan

Each sheet has its own _parse_sheet_X function. The orchestrator calls
each in turn, catching exceptions to convert into ParseWarnings.
A single bad sheet must not crash parsing of the others.
"""
from __future__ import annotations
import io
import logging
from typing import BinaryIO

import openpyxl

from app.modules.events.event_plan_excel.normalize import (
    clean_str,
    normalize_iva,
    normalize_price,
    parse_event_dates,
    parse_time_range,
    parse_topup_list,
)
from app.modules.events.event_plan_excel.schemas import (
    BarSpec,
    ContactSpec,
    ParsedEventPlan,
    ParseWarning,
    ProductSpec,
)

logger = logging.getLogger(__name__)


# Sheet name constants — exact strings as they appear in the file.
# We don't substring-match; sheet names are stable per Slesh template.
SHEET_OVERVIEW   = "1. Overview"
SHEET_PARAMETRI  = "2. Parametri Evento"
SHEET_DEVICES    = "3. Device Count"
SHEET_LISTINI    = "4. Listini Bar"
SHEET_LISTINI_F  = "4.1. Listini Food"
# SHEET_LINEUP    = "5. Lineup"   # not parsed in this version


def parse_event_plan(data: bytes | BinaryIO) -> ParsedEventPlan:
    """Parse a Slesh Event Plan .xlsx file into a typed ParsedEventPlan.

    Args:
        data: Either raw bytes or a binary file-like object.

    Returns:
        ParsedEventPlan with every field populated where possible. Any
        unparseable sheet contributes a ParseWarning instead of raising.
    """
    if isinstance(data, (bytes, bytearray)):
        stream = io.BytesIO(data)
    else:
        stream = data

    parsed = ParsedEventPlan()

    try:
        # data_only=True returns last-cached value of formula cells (Slesh
        # files don't have formulas in the sections we care about, but
        # being defensive costs us nothing).
        wb = openpyxl.load_workbook(stream, data_only=True, read_only=False)
    except Exception as exc:
        parsed.warnings.append(ParseWarning(
            sheet="(workbook)", where="open",
            message=f"could not open workbook: {type(exc).__name__}: {exc}",
        ))
        return parsed

    # Run each sheet parser independently. A failure in one sheet
    # produces a warning, never a 500.
    for sheet_name, parser_fn in [
        (SHEET_OVERVIEW,  _parse_overview),
        (SHEET_PARAMETRI, _parse_parametri),
        (SHEET_DEVICES,   _parse_devices),
        (SHEET_LISTINI,   _parse_listini_bar),
        (SHEET_LISTINI_F, _parse_listini_food),
    ]:
        if sheet_name not in wb.sheetnames:
            parsed.warnings.append(ParseWarning(
                sheet=sheet_name, where="(missing)",
                message=f"sheet not found in workbook; section skipped",
            ))
            continue
        try:
            parser_fn(wb[sheet_name], parsed)
        except Exception as exc:
            logger.exception("event_plan_excel: sheet %s parser raised", sheet_name)
            parsed.warnings.append(ParseWarning(
                sheet=sheet_name, where="(uncaught exception)",
                message=f"sheet parser failed: {type(exc).__name__}: {exc}",
            ))

    return parsed


# ─────────────────────────────────────────────────────────────────────
# Sheet 1 — Overview
# ─────────────────────────────────────────────────────────────────────
# The Overview sheet is a free-form "data dictionary" where each row
# pairs an Italian label in column B with a value in column C.
# We look up by label rather than by cell address — the template can
# shift rows between Slesh versions, but the labels are stable.
_OVERVIEW_LABEL_MAP = {
    "Nome Evento":                "event_name",
    "Orari evento":               "_orari",
    "Orario di arrivo operatori": "staff_arrival_time",
    "Date Evento":                "_dates",
    "Venue":                      "venue_name",
    "Indirizzo":                  "venue_address",
    "Capienza massima":           "capacity",
    "Presenze giornaliere previste": "expected_guests",
}


def _parse_overview(ws, parsed: ParsedEventPlan) -> None:
    """Walk column B/C as label/value pairs. Also scrape contacts from
    the 'Referenti' section below."""
    # Pass 1 — labeled fields
    for row in ws.iter_rows(min_col=2, max_col=3, values_only=True):
        label = clean_str(row[0])
        value = row[1]
        if not label or label not in _OVERVIEW_LABEL_MAP:
            continue
        target = _OVERVIEW_LABEL_MAP[label]
        if target == "_orari":
            start, end = parse_time_range(value)
            parsed.event_start_time = start
            parsed.event_end_time = end
        elif target == "_dates":
            parsed.event_dates = parse_event_dates(value)
        elif target in ("capacity", "expected_guests"):
            try:
                setattr(parsed, target, int(value) if value is not None else None)
            except (TypeError, ValueError):
                pass
        else:
            setattr(parsed, target, clean_str(value))

    # Pass 2 — contacts. The Referenti section uses rows like:
    #   B17: "Referente generale"
    #   B18: "Nome Cognome"   C18: "Alessandro Proietti"
    #   B19: "Email"          C19: "alessandro@..."
    #   B20: "Cellulare"      C20: 3383655014
    # We walk looking for "Referente ..." headers and accumulate the
    # following Nome/Email/Cellulare triples until the next header.
    current: ContactSpec | None = None
    for row in ws.iter_rows(min_col=2, max_col=3, values_only=True):
        label = clean_str(row[0])
        value = row[1]
        if not label:
            continue
        if label.lower().startswith("referente"):
            if current and (current.name or current.email or current.phone):
                parsed.contacts.append(current)
            current = ContactSpec(role=label)
            continue
        if current is None:
            continue
        if label.lower().startswith("nome"):
            current.name = clean_str(value)
        elif label.lower().startswith("email"):
            current.email = clean_str(value)
        elif label.lower().startswith("cellular") or label.lower().startswith("telefon"):
            current.phone = clean_str(value)
    if current and (current.name or current.email or current.phone):
        parsed.contacts.append(current)


# ─────────────────────────────────────────────────────────────────────
# Sheet 2 — Parametri Evento
# ─────────────────────────────────────────────────────────────────────
def _parse_parametri(ws, parsed: ParsedEventPlan) -> None:
    """Top-up denominations + refund policy. Layout is rigid enough that
    we look up by label in column A."""
    for row in ws.iter_rows(min_col=1, max_col=3, values_only=True):
        label = clean_str(row[0])
        value = row[1]
        if not label:
            continue
        low = label.lower()
        if "app operatore" in low or "app cassa" in low:
            parsed.topup_denominations_staff = parse_topup_list(value)
        elif "app utente" in low:
            parsed.topup_denominations_user = parse_topup_list(value)
        elif "credito minimo" in low:
            cents = normalize_price(value)
            if cents is not None:
                parsed.refund_min_credit_cents = cents
        elif "costo di gestione" in low:
            cents = normalize_price(value)
            if cents is not None:
                parsed.refund_fee_cents = cents


# ─────────────────────────────────────────────────────────────────────
# Sheet 3 — Device Count
# ─────────────────────────────────────────────────────────────────────
# The sheet has 4 vertically-stacked sections separated by blank rows:
#   "Punti di Ricarica"  → recharge_device_count
#   "Punti Bar"          → drink_bars
#   "Food truck"         → food_bars
#   "Altro"              → other_bars
#
# Each section starts with a header row ("NOMI PUNTI | # DISPOSITIVI |
# MENU' | DISPOSITIVO | Note"), then rows of (name, count, ...), then
# a "Tot dispositivi:" sentinel row.
_DEVICE_SECTION_HEADERS = {
    "punti di ricarica": "recharge",
    "punti bar":         "drinks",
    "food truck":        "food",
    "altro":             "other",
}


def _parse_devices(ws, parsed: ParsedEventPlan) -> None:
    section: str | None = None    # 'recharge' | 'drinks' | 'food' | 'other'
    for row in ws.iter_rows(min_col=1, max_col=5, values_only=True):
        a = clean_str(row[0])
        b = row[1]
        e = clean_str(row[4]) if len(row) > 4 else None
        if not a:
            continue
        low = a.lower()
        if low in _DEVICE_SECTION_HEADERS:
            section = _DEVICE_SECTION_HEADERS[low]
            continue
        # Skip the per-section header row (NOMI PUNTI) and the totals row
        if low.startswith("nomi punti"):
            continue
        if low.startswith("tot dispositivi"):
            continue
        # Skip section labels w/o counts (e.g. "RESP. BAR" with no number)
        try:
            count = int(b) if b is not None else 0
        except (TypeError, ValueError):
            count = 0
        if section is None or count <= 0:
            continue
        if section == "recharge":
            parsed.recharge_device_count = (parsed.recharge_device_count or 0) + count
            continue
        bar = BarSpec(
            name        = a,
            device_count= count,
            bar_type    = section if section in ("drinks", "food") else "service",
            notes       = e,
        )
        if section == "drinks":
            parsed.drink_bars.append(bar)
        elif section == "food":
            parsed.food_bars.append(bar)
        else:
            parsed.other_bars.append(bar)


# ─────────────────────────────────────────────────────────────────────
# Sheet 4 — Listini Bar
# ─────────────────────────────────────────────────────────────────────
# This sheet has TWO product layouts coexisting:
#   Rows  6-27: unified table | Nome | Negozio | Categoria | Prezzo | Iva |
#   Rows 32-88: per-vendor sub-sections, each with a "MENù PRODOTTI X"
#               header and a richer schema including Cauzione.
#
# We parse the bottom section first (richer data wins), then sweep the
# top section adding only products that didn't appear in the bottom.
def _parse_listini_bar(ws, parsed: ParsedEventPlan) -> None:
    seen_keys: set[tuple[str, str]] = set()   # (bar_name_lower, product_name_lower)

    # ── Pass 1 — per-vendor sub-sections (rows 32+)
    # State machine: "MENù PRODOTTI X" sets current_bar; then we ingest
    # rows where col A is non-empty and col D (Prezzo) is a number.
    current_bar: str | None = None
    for row in ws.iter_rows(min_row=30, min_col=1, max_col=7, values_only=True):
        a = clean_str(row[0])
        b = clean_str(row[1])
        c = clean_str(row[2])
        d = row[3]
        e = row[4]
        f = row[5] if len(row) > 5 else None
        g = row[6] if len(row) > 6 else None

        if a and a.upper().startswith("MEN"):
            # "MENù PRODOTTI MALANDRINO" / "MENù PRODOTTI COCKTAIL BAR" / etc.
            parts = a.split(" PRODOTTI ", 1)
            current_bar = clean_str(parts[1]) if len(parts) > 1 else None
            continue
        if a and a.lower() == "nome":
            continue
        # Row qualifies as a product if col A (Nome), col B (Negozio),
        # and col D (Prezzo) are present
        if not (a and b and d is not None):
            continue
        price_cents = normalize_price(d)
        iva_pct = normalize_iva(e)
        if price_cents is None or iva_pct is None:
            continue
        # Cauzione: col F is a bool (True/False), col G is the amount.
        # Some rows have F=False with G=numeric (the deposit applies but
        # isn't 'charged' — Slesh convention). We always treat G as the
        # deposit amount when present.
        cauzione_cents = normalize_price(g) or 0
        key = (b.lower(), a.lower())
        if key in seen_keys:
            continue
        seen_keys.add(key)
        parsed.products.append(ProductSpec(
            name           = a,
            bar_name       = b,
            category       = c or "",
            price_cents    = price_cents,
            iva_pct        = iva_pct,
            cauzione_cents = cauzione_cents,
        ))

    # ── Pass 2 — top unified table (rows 6-27), only add unseen items
    for row in ws.iter_rows(min_row=6, max_row=28, min_col=1, max_col=5, values_only=True):
        a = clean_str(row[0])
        b = clean_str(row[1])
        c = clean_str(row[2])
        d = row[3]
        e = row[4]
        if not (a and b and d is not None):
            continue
        if a.lower() == "nome":   # the header row
            continue
        price_cents = normalize_price(d)
        iva_pct = normalize_iva(e)
        if price_cents is None or iva_pct is None:
            continue
        key = (b.lower(), a.lower())
        if key in seen_keys:
            continue
        seen_keys.add(key)
        parsed.products.append(ProductSpec(
            name           = a,
            bar_name       = b,
            category       = c or "",
            price_cents    = price_cents,
            iva_pct        = iva_pct,
            cauzione_cents = 0,
        ))


# ─────────────────────────────────────────────────────────────────────
# Sheet 4.1 — Listini Food (per-vendor side-by-side tables)
# ─────────────────────────────────────────────────────────────────────
# Layout: two adjacent table blocks per row range.
#   Cols A-C: Vendor 1 (Nome | Prezzo | Iva)
#   Cols E-G: Vendor 2 (Nome | Prezzo | Iva)
# Vendor name lives in a "LISTINO PREZZI - X" header above each block.
def _parse_listini_food(ws, parsed: ParsedEventPlan) -> None:
    seen_keys: set[tuple[str, str]] = set(
        (p.bar_name.lower(), p.name.lower()) for p in parsed.products
    )

    # Scan for vendor headers and remember their column range
    # (left block A-C, right block E-G).
    blocks: list[tuple[str, int, int, int]] = []  # (vendor, header_row, name_col, price_col, iva_col)

    # Header detection: any cell whose value starts with "LISTINO PREZZI"
    # marks a vendor block starting at the row below.
    for row in ws.iter_rows(min_col=1, max_col=7, values_only=False):
        for cell in row:
            v = clean_str(cell.value)
            if not v or not v.upper().startswith("LISTINO PREZZI"):
                continue
            # vendor is what follows the dash (e.g. "MALANDRINO")
            parts = v.split("-", 1)
            vendor = clean_str(parts[1]) if len(parts) > 1 else v
            if not vendor:
                continue
            # Block geometry: data starts 2 rows below (header row + "Nome | Prezzo | Iva" row + data)
            # Columns depend on which cell we found this in.
            col = cell.column   # 1-indexed
            blocks.append((vendor, cell.row, col, col + 1, col + 2))

    # For each block, read down until rows go empty.
    for vendor, header_row, name_col, price_col, iva_col in blocks:
        start_row = header_row + 2   # skip the "Nome | Prezzo | Iva" sub-header
        for r in range(start_row, start_row + 20):  # max 20 items per vendor — generous
            name = clean_str(ws.cell(row=r, column=name_col).value)
            price_raw = ws.cell(row=r, column=price_col).value
            iva_raw = ws.cell(row=r, column=iva_col).value
            if not name:
                continue
            price_cents = normalize_price(price_raw)
            iva_pct = normalize_iva(iva_raw)
            if price_cents is None or iva_pct is None:
                continue
            key = (vendor.lower(), name.lower())
            if key in seen_keys:
                continue
            seen_keys.add(key)
            parsed.products.append(ProductSpec(
                name           = name,
                bar_name       = vendor,
                category       = "Panineria",   # Sheet 4.1 is food-only
                price_cents    = price_cents,
                iva_pct        = iva_pct,
                cauzione_cents = 0,
            ))
