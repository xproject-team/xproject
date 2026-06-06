"""Seed a dedicated tenant for simulator dry-runs.

Creates a fully isolated sim environment:
  - Tenant:  "Sim 2025-06-15"
  - Venue:   "Sundance 2025 Simulation"
  - Owner:   sim-owner@noma-sim.test  (password: simulator)
  - Event:   "Sim Sundance 2025-06-15" — LIVE, 12h window from now
  - 8 bars:  matching Slesh 2025 shop names verbatim
             (Beer Bar, Cocktail Bar, Malandrino, Focacceria,
              Figo, La Nina, Gelateria, Guardaroba)
  - 32 products: from the 2025 Prodotti-categoria XLSX, upserted
                 against any existing tenant products by name
  - bar_stock: 1000 units allocated per (bar, drink),
               500 per (bar, food)

Run with:
    python -m app.scripts.seed_sim_event

Idempotent: re-running on the same DB is a no-op apart from
returning the existing IDs. Drop via SQL:
    DELETE FROM tenants WHERE slug = \'sim-2025-06-15\';
(cascades through all child tables)

S5 of docs/e2e-validation-design.md.
"""
import asyncio
import sys
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import AsyncSessionLocal
from app.core.security import hash_password
from app.modules.auth.models import Tenant, User, UserRole, UserRoleAssignment
from app.modules.bar_stock.models import BarStock
from app.modules.bars.models import Bar
from app.modules.events.models import Event, EventStatus
from app.modules.products.models import (
    Product, ProductCategory, ProductType, ProductUnit,
)
from app.modules.venues.models import Venue


# ─── Constants ────────────────────────────────────────────────────────

TENANT_SLUG  = "sim-2025-06-15"
TENANT_NAME  = "Sim 2025-06-15"
OWNER_EMAIL  = "sim-owner@noma-sim.test"
OWNER_PASS   = "simulator"
VENUE_NAME   = "Sundance 2025 Simulation"
EVENT_NAME   = "Sim Sundance 2025-06-15"

# Bar names — UNION of every shop name observed across the 5
# historical events (data_DD_MM_2025/2-ordini_bracciali/*.xlsx).
# Per Omar's product policy, every Sundance event has a different
# bar lineup; this list is the COMPLETE catalog so the sim tenant
# can run any historical event without skip-unknown-bar errors.
SIM_BAR_NAMES = [
    "Ape Magna",
    "Beer Bar",
    "Cocktail Bar",
    "Figo",
    "Focacceria",
    "Gelateria",
    "Guardaroba",
    "La Nina",
    "Malandrino",
    "Merch",
    "Pret a Polpett",
    "Stravizio",
    "Totò",
    "Twist&Chips",
]

# Slesh category one-hot -> our ProductType + reasonable defaults
def _map_category(bev: int, food: int, guard: int) -> tuple[ProductType, ProductCategory | None, ProductUnit]:
    """Map Slesh one-hot to (ProductType, ProductCategory, ProductUnit).

    Heuristic: drinks land as SOFT_DRINK unless clearly a cocktail
    by name; food + supplies get no category. The simulator
    doesn\'t exercise tier_rank logic so basic mappings are fine.
    """
    if guard:
        return ProductType.SUPPLY, None, ProductUnit.PIECE
    if food:
        return ProductType.FOOD, None, ProductUnit.PIECE
    return ProductType.DRINK, ProductCategory.SOFT_DRINK, ProductUnit.GLASS


def _refine_drink_category(name: str, default: ProductCategory) -> ProductCategory | None:
    """Light-touch refinement based on product name."""
    lname = name.lower()
    if "cocktail" in lname:
        if "premium" in lname or "super" in lname:
            return ProductCategory.PREMIUM_COCKTAIL
        return ProductCategory.BASIC_COCKTAIL
    if "vino" in lname or "bolle" in lname:
        return ProductCategory.WINE_SPARKLING
    if "raffo" in lname or "nastro" in lname:
        return ProductCategory.BEER_BOTTLE
    if "sprtiz" in lname or "spritz" in lname:
        return ProductCategory.BASIC_COCKTAIL
    return default


# Root of all 5 historical events. Each event has its own
# Prodotti-categoria XLSX with its own menu; the seed unions
# all of them.
HISTORICAL_DATA_ROOT = Path.home() / "Desktop" / "2025"


# ─── Helpers ──────────────────────────────────────────────────────────

async def _get_or_create_tenant(db: AsyncSession) -> Tenant:
    r = await db.execute(select(Tenant).where(Tenant.slug == TENANT_SLUG))
    t = r.scalar_one_or_none()
    if t is not None:
        print(f"  tenant exists (id={t.id})")
        return t
    t = Tenant(name=TENANT_NAME, slug=TENANT_SLUG)
    db.add(t); await db.flush()
    print(f"  + created tenant id={t.id}")
    return t


async def _get_or_create_owner(db: AsyncSession, tenant_id) -> User:
    r = await db.execute(
        select(User).where(User.tenant_id == tenant_id, User.email == OWNER_EMAIL)
    )
    u = r.scalar_one_or_none()
    if u is None:
        u = User(
            tenant_id=tenant_id,
            email=OWNER_EMAIL,
            hashed_password=hash_password(OWNER_PASS),
            full_name="Sim Owner",
            role=UserRole.OWNER,
            is_active=True,
        )
        db.add(u); await db.flush()
        print(f"  + created owner id={u.id}")
    else:
        print(f"  owner exists (id={u.id})")
    # Login authorizes against the user_roles TABLE, not the legacy User.role
    # column. The p1 migration backfilled pre-existing users; anyone seeded
    # afterwards needs the grant explicitly or they 403 on a role-scoped login.
    existing = await db.execute(
        select(UserRoleAssignment).where(
            UserRoleAssignment.user_id == u.id,
            UserRoleAssignment.role == UserRole.OWNER,
        )
    )
    if existing.scalar_one_or_none() is None:
        db.add(UserRoleAssignment(user_id=u.id, role=UserRole.OWNER))
        await db.flush()
        print("  + granted OWNER authorization row")
    return u


async def _get_or_create_venue(db: AsyncSession, tenant_id) -> Venue:
    r = await db.execute(
        select(Venue).where(Venue.tenant_id == tenant_id, Venue.name == VENUE_NAME)
    )
    v = r.scalar_one_or_none()
    if v is not None:
        return v
    v = Venue(tenant_id=tenant_id, name=VENUE_NAME, address="Simulator",
              capacity=2000)
    db.add(v); await db.flush()
    return v


async def _get_or_create_event(db: AsyncSession, tenant_id, venue_id) -> Event:
    r = await db.execute(
        select(Event).where(Event.tenant_id == tenant_id, Event.name == EVENT_NAME)
    )
    e = r.scalar_one_or_none()
    if e is not None:
        print(f"  event exists (id={e.id}, status={e.status.value})")
        return e
    now = datetime.now(timezone.utc)
    e = Event(
        tenant_id=tenant_id,
        venue_id=venue_id,
        name=EVENT_NAME,
        scheduled_at=now - timedelta(hours=1),
        scheduled_end_at=now + timedelta(hours=12),
        # Status starts as DRAFT — caller may flip to LIVE.
        status=EventStatus.DRAFT,
        expected_guest_count=500,
        version=1,
    )
    db.add(e); await db.flush()
    print(f"  + created event id={e.id} (DRAFT)")
    return e


async def _get_or_create_bars(db: AsyncSession, tenant_id, event_id) -> list[Bar]:
    bars: list[Bar] = []
    for name in SIM_BAR_NAMES:
        r = await db.execute(
            select(Bar).where(
                Bar.tenant_id == tenant_id,
                Bar.event_id == event_id,
                Bar.name == name,
            )
        )
        b = r.scalar_one_or_none()
        if b is None:
            b = Bar(
                tenant_id=tenant_id, event_id=event_id, name=name,
                bar_type="food" if name in (
                    "Focacceria", "Gelateria", "Malandrino",
                    "Pret a Polpett", "Twist&Chips", "Totò", "Ape Magna",
                ) else ("merch" if name == "Merch" else "drinks"),
                is_active=True,
            )
            db.add(b); await db.flush()
            print(f"  + bar {name!r:20s} id={b.id}")
        bars.append(b)
    return bars


async def _upsert_products(db: AsyncSession, tenant_id) -> dict[str, Product]:
    """Scan ALL 5 historical events\' Prodotti-categoria XLSX files,
    take the union, upsert each product by name.

    Different events have different menus (Omar product policy).
    The sim tenant gets the complete catalog so any event can be
    replayed without skipping rows. Conflicts (same product name,
    different prices across events) resolve last-write-wins by
    sorted event directory name — deterministic and reproducible.
    """
    rows_unioned: dict[str, tuple] = {}  # name -> (name, prezzo, totale, bev, food, guard)
    event_dirs = sorted(HISTORICAL_DATA_ROOT.glob("data_*_2025"))
    if not event_dirs:
        print(f"  ❌ no event dirs found under {HISTORICAL_DATA_ROOT}")
        sys.exit(1)

    for event_dir in event_dirs:
        # Look for the Prodotti file that has category columns.
        # Naming varies:
        #   15_06_2025 → "Prodotti-categoria-SUNDANCE..." (cols: name,prezzo,totale,beverage,Food,Guardaroba)
        #   13_07_2025 → "Prodotti-Sundance...-1775123995664.xlsx" (cols: name,prezzo,totale,beverage,Food)
        # The OTHER Prodotti-*.xlsx in each dir has 4 cols
        # (name, Totale prodotti, Totale fatturato, Totale Consumabili)
        # and no category info — skip that one.
        picked = None
        for f in sorted((event_dir / "3-prodotti").glob("Prodotti-*.xlsx")):
            wb = openpyxl.load_workbook(f, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            # Heuristic: file has "Prezzo" (price) col AND at least one
            # category one-hot column.
            has_price = "Prezzo" in headers
            has_category = any(c in headers for c in ("beverage", "Food", "Guardaroba"))
            wb.close()
            if has_price and has_category:
                picked = f
                break  # first match by sorted name wins
        if picked is None:
            print(f"  ⚠️  no category-aware Prodotti file in {event_dir.name}")
            continue

        # Re-open and read all rows; locate columns by header (not position)
        wb = openpyxl.load_workbook(picked, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col = {h: headers.index(h) for h in headers if h}
        idx_name  = col.get("Prodotto", 0)
        idx_prezzo = col.get("Prezzo", 1)
        idx_totale = col.get("Totale", 2)
        idx_bev   = col.get("beverage")
        idx_food  = col.get("Food")
        idx_guard = col.get("Guardaroba")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[idx_name] is None:
                continue
            name = str(row[idx_name]).strip()
            prezzo = row[idx_prezzo] if idx_prezzo is not None else None
            totale = row[idx_totale] if idx_totale is not None else None
            bev   = row[idx_bev]   if idx_bev   is not None else 0
            food  = row[idx_food]  if idx_food  is not None else 0
            guard = row[idx_guard] if idx_guard is not None else 0
            # Build a normalized 6-tuple matching the original shape
            rows_unioned[name] = (name, prezzo, totale, bev, food, guard)
        wb.close()
    print(f"  scanned {len(event_dirs)} events; union of products: {len(rows_unioned)}")

    products_by_name: dict[str, Product] = {}
    for row in rows_unioned.values():
        name, prezzo, totale, bev, food, guard = row
        if name is None:
            continue
        name = str(name).strip()
        prezzo_cents = int(round(float(prezzo) * 100)) if prezzo is not None else None

        ptype, default_cat, punit = _map_category(int(bev or 0), int(food or 0), int(guard or 0))
        if ptype == ProductType.DRINK:
            category = _refine_drink_category(name, default_cat)
        else:
            category = None

        # Check existing
        r = await db.execute(
            select(Product).where(
                Product.tenant_id == tenant_id,
                Product.name == name,
            )
        )
        p = r.scalar_one_or_none()
        if p is None:
            p = Product(
                tenant_id=tenant_id,
                name=name,
                product_type=ptype,
                category=category,
                unit=punit,
                default_price_cents=prezzo_cents,
                tier_rank=2 if ptype == ProductType.DRINK else None,
            )
            db.add(p); await db.flush()
        products_by_name[name] = p

    print(f"  + products upserted: {len(products_by_name)}")
    return products_by_name


async def _seed_bar_stock(db: AsyncSession, tenant_id, event_id,
                          bars: list[Bar], products: dict[str, Product]) -> int:
    """Seed bar_stock with generous quantities so depletion doesn\'t
    fire during the smoke run. Tighten later if we want alerts."""
    count = 0
    for bar in bars:
        for name, prod in products.items():
            r = await db.execute(
                select(BarStock).where(
                    BarStock.bar_id == bar.id,
                    BarStock.product_id == prod.id,
                )
            )
            if r.scalar_one_or_none() is not None:
                continue
            alloc = Decimal("1000") if prod.product_type == ProductType.DRINK \
                    else Decimal("500")
            bs = BarStock(
                tenant_id=tenant_id,
                event_id=event_id,
                bar_id=bar.id,
                product_id=prod.id,
                allocated_qty=alloc,
                current_qty=alloc,
                returned_qty=Decimal("0"),
            )
            db.add(bs)
            count += 1
    await db.flush()
    print(f"  + bar_stock rows seeded: {count}")
    return count


# ─── Main ─────────────────────────────────────────────────────────────

async def seed_sim_event() -> None:
    print("Seeding simulator tenant + event...")
    async with AsyncSessionLocal() as db:
        tenant = await _get_or_create_tenant(db)
        owner  = await _get_or_create_owner(db, tenant.id)
        venue  = await _get_or_create_venue(db, tenant.id)
        event  = await _get_or_create_event(db, tenant.id, venue.id)
        bars   = await _get_or_create_bars(db, tenant.id, event.id)
        products = await _upsert_products(db, tenant.id)
        await _seed_bar_stock(db, tenant.id, event.id, bars, products)
        await db.commit()

        print()
        print("═══ Summary ═══")
        print(f"  tenant_id:  {tenant.id}")
        print(f"  event_id:   {event.id}  (status={event.status.value})")
        print(f"  owner:      {OWNER_EMAIL} / {OWNER_PASS}")
        print(f"  bars:       {len(bars)}")
        print(f"  products:   {len(products)}")
        print()
        print("Next: promote the event to LIVE before running the simulator:")
        print(f"  psql xproject_dev -c \"")
        print(f"    UPDATE events SET status=\'LIVE\', started_at=NOW() ")
        print(f"    WHERE id=\'{event.id}\';\"")


if __name__ == "__main__":
    asyncio.run(seed_sim_event())
